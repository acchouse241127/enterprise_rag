"""Excel parser implementation with smart table detection.

Author: C2
Date: 2026-03-06 (Updated for V2.1 with smart table detection)
"""

from pathlib import Path

from .base import BaseDocumentParser
from .models import ContentType, ParsedContent


# Threshold for "small table" vs "large table"
# Small tables (< threshold) are kept as complete Markdown
# Large tables (>= threshold) are split by rows with header prefix
SMALL_TABLE_ROW_THRESHOLD = 20


class ExcelDocumentParser(BaseDocumentParser):
    """Parser for Excel documents with smart table detection.

    Strategy (Option C):
    - Small tables (<20 rows): Output complete Markdown table
    - Large tables (>=20 rows): Output header description + row-by-row chunks
    """

    def __init__(self, row_threshold: int = SMALL_TABLE_ROW_THRESHOLD) -> None:
        """Initialize parser with row threshold.

        Args:
            row_threshold: Number of rows to distinguish small/large tables
        """
        self.row_threshold = row_threshold

    def parse(self, file_path: Path) -> list[ParsedContent]:
        """Parse Excel file and return structured content."""
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(filename=str(file_path), data_only=True)
        contents: list[ParsedContent] = []

        for sheet in wb.worksheets:
            sheet_contents = self._parse_sheet(sheet)
            contents.extend(sheet_contents)

        return contents

    def _parse_sheet(self, sheet) -> list[ParsedContent]:
        """Parse a single worksheet.

        Args:
            sheet: openpyxl worksheet object

        Returns:
            List of ParsedContent for this sheet
        """
        # Get all rows
        rows_data = list(sheet.iter_rows(values_only=True))
        if not rows_data:
            return []

        # Find header row (first non-empty row)
        header_row = None
        header_idx = 0
        for idx, row in enumerate(rows_data):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                header_row = row
                header_idx = idx
                break

        if header_row is None:
            return []

        # Extract column names
        headers = []
        for cell in header_row:
            if cell is not None and str(cell).strip():
                headers.append(str(cell).strip())
            else:
                headers.append("")

        # Get data rows
        data_rows = rows_data[header_idx + 1:]
        valid_data_rows = [
            row for row in data_rows
            if any(cell is not None and str(cell).strip() for cell in row)
        ]

        if not valid_data_rows:
            return []

        # Decide strategy based on row count
        if len(valid_data_rows) < self.row_threshold:
            # Small table: output complete Markdown
            return self._build_small_table(sheet.title, headers, valid_data_rows)
        else:
            # Large table: output header description + row chunks
            return self._build_large_table(sheet.title, headers, valid_data_rows)

    def _build_small_table(
        self,
        sheet_name: str,
        headers: list[str],
        data_rows: list,
    ) -> list[ParsedContent]:
        """Build complete Markdown table for small tables."""
        # Build Markdown table
        md_lines = [f"[工作表: {sheet_name}]", ""]
        md_lines.append("| " + " | ".join(headers) + " |")
        md_lines.append("|" + "|".join(["---"] * len(headers)) + "|")

        for row in data_rows:
            cells = []
            for i, cell in enumerate(row):
                if i < len(headers):
                    cells.append(str(cell) if cell is not None else "")
                else:
                    cells.append(str(cell) if cell is not None else "")
            md_lines.append("| " + " | ".join(cells) + " |")

        table_md = "\n".join(md_lines)

        return [
            ParsedContent(
                content_type=ContentType.TABLE,
                text=table_md,
                metadata={
                    "sheet_name": sheet_name,
                    "table_markdown": table_md,
                    "row_count": len(data_rows),
                },
            )
        ]

    def _build_large_table(
        self,
        sheet_name: str,
        headers: list[str],
        data_rows: list,
    ) -> list[ParsedContent]:
        """Build header description + row chunks for large tables."""
        contents: list[ParsedContent] = []

        # Header description
        header_desc = f"[工作表: {sheet_name} - 共 {len(data_rows)} 行数据]\n"
        header_desc += f"表头: {' | '.join(h for h in headers if h)}"

        contents.append(
            ParsedContent(
                content_type=ContentType.TABLE,
                text=header_desc,
                metadata={
                    "sheet_name": sheet_name,
                    "is_header": True,
                    "row_count": len(data_rows),
                },
            )
        )

        # Row-by-row chunks
        for row in data_rows:
            row_parts = []
            for i, cell in enumerate(row):
                if cell is None or not str(cell).strip():
                    continue
                col_name = headers[i] if i < len(headers) and headers[i] else f"列{i+1}"
                row_parts.append(f"{col_name}: {str(cell).strip()}")

            if row_parts:
                row_text = " | ".join(row_parts)
                contents.append(
                    ParsedContent(
                        content_type=ContentType.TABLE,
                        text=row_text,
                        metadata={
                            "sheet_name": sheet_name,
                            "is_row": True,
                        },
                    )
                )

        return contents
